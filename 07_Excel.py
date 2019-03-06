
# 【1】==导入包
from openpyxl import Workbook;
from openpyxl.drawing.image import Image;
import PIL
# 【2】==创建一个Excel表格
wb = Workbook();
# 【4】==默认创建一个：Sheet选项卡
# 添加一个选项卡
wb.create_sheet("第一个选项卡",0);
wb.create_sheet("第二个选项卡",1);
# 【5】删除一个选项卡
wb.remove_sheet(wb["Sheet"]);
# 【6】获取想要操作的选项卡 -active 获取当前活跃状态的选项卡
ws = wb.active;
# 【7】通过选项卡名称进行选择
ws1 = wb["第二个选项卡"];
# 【8】赋值操作
ws1["A1"] = "姓名";
ws1["B1"] = "年龄"
# 【9】根据行列进行控制
ws1.cell(row=2,column=1,value = "狄仁杰");
# 【10】整行输入
ws1.append(["姓名","年龄","性别"])
array_01 = [["狄仁杰1","181","男"],["狄仁杰2","182","男"],["狄仁杰3","183","男"]];
for i in range(0,len(array_01)):
    ws1.append(array_01[i]);
    pass
# 【11】插入一个张图片
ws.column_dimensions["A"].width = 100;
ws.row_dimensions[1].height = 100;
# 需要一张图片
img= Image("1.jpg");
img.width = 100;
img.height = 100;
ws.add_image(img,"A1");


print(ws)


# 【3】==存储
wb.save("2019年3月6日.xlsx")